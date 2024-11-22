from openpyxl import load_workbook
from openpyxl.chart import PieChart, Reference

# Открываем созданный файл Excel
wb = load_workbook("Зарплаты.xlsx")
ws = wb.active

# Считываем данные
rows = list(ws.iter_rows(min_row=2, values_only=True))
data = [row for row in rows if row[0] and "Итого" not in row[1]]

# Суммируем зарплаты по отделам
total_salary_by_dept = {}

for row in data:
    dept = row[2]  # Название отдела
    salary = row[5]  # Сумма зарплаты
    if dept not in total_salary_by_dept:
        total_salary_by_dept[dept] = 0
    total_salary_by_dept[dept] += salary

# Вставляем данные для диаграммы в Excel
chart_data = []
for dept, total_salary in total_salary_by_dept.items():
    chart_data.append([dept, total_salary])

# Определяем, где начать вставку данных для диаграммы
# Считаем количество строк, занятых таблицей, чтобы избежать их перекрытия
start_row = len(rows) + 2  # Место, где начнем вставку данных для диаграммы

# Добавляем данные на лист Excel
ws.append(["Отдел", "Сумма зарплаты"])  # Заголовки для данных диаграммы
for row in chart_data:
    ws.append(row)

# Создаем круговую диаграмму
chart = PieChart()
data = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + len(chart_data) - 1)
categories = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + len(chart_data) - 1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)
chart.title = "Распределение зарплаты по отделам"

# Вставляем диаграмму в свободную область справа от таблицы
chart_position = f"F{start_row}"  # Позиция диаграммы справа от таблицы
ws.add_chart(chart, chart_position)

# Сохраняем файл
wb.save("Зарплаты_с_диаграммой.xlsx")
