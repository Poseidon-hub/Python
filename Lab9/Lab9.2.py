from openpyxl import load_workbook

# Открываем созданный файл
wb = load_workbook("Зарплаты.xlsx")
ws = wb.active

# Считываем данные
rows = list(ws.iter_rows(min_row=2, values_only=True))
data = [row for row in rows if row[0] and "Итого" not in row[1]]

# Вычисления
max_salary = max(data, key=lambda x: x[5])
min_salary = min(data, key=lambda x: x[5])
average_salary_by_dept = {}

for row in data:
    dept = row[2]
    if dept not in average_salary_by_dept:
        average_salary_by_dept[dept] = []
    average_salary_by_dept[dept].append(row[5])

average_salary_by_dept = {k: sum(v)/len(v) for k, v in average_salary_by_dept.items()}

# Вывод
print(f"Максимальная зарплата: {max_salary[1]} - {max_salary[5]} руб.")
print(f"Минимальная зарплата: {min_salary[1]} - {min_salary[5]} руб.")
print("Средняя зарплата по отделам:")
for dept, avg in average_salary_by_dept.items():
    print(f"{dept}: {round(avg, 2)} руб.")

# на 14 строке и втором столбце пустые значения